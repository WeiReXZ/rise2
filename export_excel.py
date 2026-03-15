"""Выгрузка участников и визитов без регистрации в Excel."""
import io
from typing import Any, List

import openpyxl
from openpyxl.styles import Font, Alignment
from openpyxl.utils import get_column_letter


def build_excel_bytes(
    participants: List[dict],
    visitors: List[dict] = None,
) -> io.BytesIO:
    wb = openpyxl.Workbook()
    visitors = visitors or []

    # Лист 1: Участники
    ws1 = wb.active
    ws1.title = "Участники"
    headers1 = [
        "№ участника",
        "Telegram ID",
        "Username",
        "Имя",
        "Возраст",
        "Род деятельности",
        "Цель",
        "Телефон",
        "Дата регистрации",
        "Дата выхода",
        "Источник",
    ]
    key_order1 = [
        "participant_number",
        "telegram_id",
        "username",
        "name",
        "age",
        "occupation",
        "goal",
        "phone",
        "created_at",
        "left_at",
        "source",
    ]
    for col, h in enumerate(headers1, 1):
        cell = ws1.cell(row=1, column=col, value=h)
        cell.font = Font(bold=True)
        cell.alignment = Alignment(horizontal="center", wrap_text=True)
    for row_idx, row_data in enumerate(participants, 2):
        for col_idx, key in enumerate(key_order1, 1):
            ws1.cell(row=row_idx, column=col_idx, value=row_data.get(key, ""))
    for col in range(1, len(headers1) + 1):
        ws1.column_dimensions[get_column_letter(col)].width = 18

    # Лист 2: Визиты без регистрации
    ws2 = wb.create_sheet("Визиты без регистрации")
    headers2 = ["Telegram ID", "Источник", "Дата первого визита"]
    key_order2 = ["telegram_id", "source", "first_seen_at"]
    for col, h in enumerate(headers2, 1):
        cell = ws2.cell(row=1, column=col, value=h)
        cell.font = Font(bold=True)
        cell.alignment = Alignment(horizontal="center", wrap_text=True)
    for row_idx, row_data in enumerate(visitors, 2):
        for col_idx, key in enumerate(key_order2, 1):
            val = row_data.get(key, "")
            if key == "first_seen_at" and val:
                val = (str(val))[:19].replace("T", " ") if len(str(val)) >= 19 else val
            ws2.cell(row=row_idx, column=col_idx, value=val)
    for col in range(1, len(headers2) + 1):
        ws2.column_dimensions[get_column_letter(col)].width = 18

    buf = io.BytesIO()
    wb.save(buf)
    buf.seek(0)
    return buf
