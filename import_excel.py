"""Импорт из Excel в том же формате, что export_excel.build_excel_bytes."""
import io
from datetime import datetime, timezone
from typing import Any, List, Tuple

import openpyxl

SHEET_PARTICIPANTS = "Участники"
SHEET_VISITS = "Визиты без регистрации"

HEADERS_PARTICIPANTS = [
    "№ участника",
    "Telegram ID",
    "Username",
    "Имя",
    "Возраст",
    "Род деятельности",
    "Цель",
    "Телефон",
    "Дата регистрации",
    "Дата выхода",
    "Источник",
]

HEADERS_VISITS = ["Telegram ID", "Источник", "Дата первого визита"]


def _dt_to_iso(dt: datetime) -> str:
    if dt.tzinfo is None:
        dt = dt.replace(tzinfo=timezone.utc)
    return dt.astimezone(timezone.utc).replace(microsecond=0).isoformat().replace("+00:00", "Z")


def _cell_text(v: Any) -> str:
    if v is None:
        return ""
    if isinstance(v, datetime):
        return _dt_to_iso(v)
    return str(v).strip()


def _cell_int(v: Any) -> int:
    if v is None:
        raise ValueError("пустое значение")
    if isinstance(v, bool):
        raise ValueError("неверный тип")
    if isinstance(v, float):
        if not v.is_integer():
            raise ValueError(f"ожидалось целое: {v}")
        return int(v)
    if isinstance(v, int):
        return v
    s = str(v).strip()
    if not s:
        raise ValueError("пустое значение")
    if "." in s or "e" in s.lower():
        f = float(s)
        if not f.is_integer():
            raise ValueError(f"ожидалось целое: {v}")
        return int(f)
    return int(s)


def parse_excel_bytes(data: bytes) -> Tuple[List[dict], List[dict]]:
    """
    Парсит .xlsx из выгрузки «Выгрузить Excel».
    Возвращает (участники, визиты). ValueError при неверном формате.
    """
    wb = None
    try:
        wb = openpyxl.load_workbook(io.BytesIO(data), data_only=True)
    except Exception as e:
        raise ValueError(f"Не удалось открыть Excel: {e}") from e

    try:
        if SHEET_PARTICIPANTS not in wb.sheetnames:
            raise ValueError(
                f"Нет листа «{SHEET_PARTICIPANTS}». Нужны те же имена листов, что в выгрузке."
            )
        if SHEET_VISITS not in wb.sheetnames:
            raise ValueError(f"Нет листа «{SHEET_VISITS}».")

        ws = wb[SHEET_PARTICIPANTS]
        rows = list(ws.iter_rows(values_only=True))
        if not rows:
            raise ValueError("Лист «Участники» пуст.")
        headers = ["" if h is None else str(h).strip() for h in rows[0]]
        if headers != HEADERS_PARTICIPANTS:
            raise ValueError(
                "Заголовки листа «Участники» не совпадают с выгрузкой. "
                "Скачайте «Выгрузить Excel» и не меняйте первую строку."
            )

        participants: List[dict] = []
        seen_tg: set = set()
        seen_num: set = set()
        for i, row in enumerate(rows[1:], start=2):
            if not row:
                continue
            if all(x is None or (isinstance(x, str) and not x.strip()) for x in row):
                continue
            row = list(row)
            while len(row) < 11:
                row.append(None)
            row = row[:11]
            try:
                pn = _cell_int(row[0])
                tg = _cell_int(row[1])
            except ValueError as e:
                raise ValueError(f"Участники, строка {i}: {e}") from e
            if tg in seen_tg:
                raise ValueError(f"Дубликат Telegram ID {tg} в участниках.")
            seen_tg.add(tg)
            if pn in seen_num:
                raise ValueError(f"Дубликат № участника {pn}.")
            seen_num.add(pn)

            name = _cell_text(row[3])
            age = _cell_text(row[4])
            occ = _cell_text(row[5])
            goal = _cell_text(row[6])
            phone = _cell_text(row[7])
            created_raw = row[8]
            if created_raw is None or (isinstance(created_raw, str) and not created_raw.strip()):
                raise ValueError(f"Участники, строка {i}: нужна дата регистрации.")
            created_at = _dt_to_iso(created_raw) if isinstance(created_raw, datetime) else str(created_raw).strip()
            if not all([name, age, occ, goal, phone, created_at]):
                raise ValueError(
                    f"Участники, строка {i}: заполните имя, возраст, деятельность, цель, телефон, дату регистрации."
                )

            left_raw = row[9]
            left_at = None
            if left_raw is not None and str(left_raw).strip() != "":
                left_at = _dt_to_iso(left_raw) if isinstance(left_raw, datetime) else str(left_raw).strip()

            src = _cell_text(row[10])
            participants.append(
                {
                    "participant_number": pn,
                    "telegram_id": tg,
                    "username": _cell_text(row[2]),
                    "name": name,
                    "age": age,
                    "occupation": occ,
                    "goal": goal,
                    "phone": phone,
                    "created_at": created_at,
                    "left_at": left_at,
                    "source": src or None,
                }
            )

        ws2 = wb[SHEET_VISITS]
        rows2 = list(ws2.iter_rows(values_only=True))
        visitors: List[dict] = []
        if rows2:
            h2 = ["" if h is None else str(h).strip() for h in rows2[0]]
            if h2 != HEADERS_VISITS:
                raise ValueError("Заголовки листа «Визиты без регистрации» не совпадают с выгрузкой.")
            seen_v: set = set()
            for i, row in enumerate(rows2[1:], start=2):
                if not row:
                    continue
                if all(x is None or (isinstance(x, str) and not str(x).strip()) for x in row):
                    continue
                row = list(row)
                while len(row) < 3:
                    row.append(None)
                row = row[:3]
                try:
                    tg = _cell_int(row[0])
                except ValueError as e:
                    raise ValueError(f"Визиты, строка {i}: {e}") from e
                if tg in seen_v:
                    raise ValueError(f"Дубликат Telegram ID {tg} на листе визитов.")
                seen_v.add(tg)
                src = _cell_text(row[1])
                fs_raw = row[2]
                if fs_raw is None or str(fs_raw).strip() == "":
                    raise ValueError(f"Визиты, строка {i}: нужна дата первого визита.")
                first_seen = (
                    _dt_to_iso(fs_raw) if isinstance(fs_raw, datetime) else str(fs_raw).strip()
                )
                visitors.append(
                    {
                        "telegram_id": tg,
                        "source": src or None,
                        "first_seen_at": first_seen,
                    }
                )

        return participants, visitors
    finally:
        if wb is not None:
            wb.close()
